from db import Database
from typing import List, Dict, Any, Optional
import json
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

class Backend:
    def __init__(self):
        self.db = Database()
    
    def register_user(self, user_id: int, username: str, first_name: str, last_name: str, phone: str = None, role: str = 'user'):
        existing_user = self.db.get_user(user_id)
        if existing_user:
            self.db.execute(
                "UPDATE users SET username = ?, first_name = ?, last_name = ?, phone = ? WHERE user_id = ?",
                (username, first_name, last_name, phone, user_id)
            )
        else:
            self.db.add_user(user_id, username, first_name, last_name, phone, role)

    def update_user_phone(self, user_id: int, phone: str):
        self.db.update_user_phone(user_id, phone)
    
    def get_user_role(self, user_id: int) -> Optional[str]:
        user = self.db.get_user(user_id)
        return user.get('role') if user else None
    

    def get_user_by_phone(self, phone: str) -> Optional[Dict]:
        return self.db.get_user_by_phone(phone)
    
    def remove_driver(self, username: str) -> bool:
        user = self.db.fetch_one("SELECT user_id FROM users WHERE username = ?", (username,))
        if not user:
            return False
        
        driver_user_id = user['user_id']
        
        self.db.execute("DELETE FROM order_responses WHERE driver_id = ?", (driver_user_id,))
        self.db.execute("DELETE FROM driver_offers WHERE driver_id = ?", (driver_user_id,))
        self.db.execute("DELETE FROM drivers WHERE user_id = ?", (driver_user_id,))
        self.db.execute("UPDATE users SET role = 'user' WHERE user_id = ?", (driver_user_id,))
        
        return True
    
    def get_driver_info(self, user_id: int) -> Optional[Dict]:
        driver = self.db.get_driver(user_id)
        if driver and driver.get('group_id'):
            group = self.db.get_group(driver['group_id'])
            if group:
                driver['group_name'] = group['group_name']
        return driver
    
    def get_driver_by_username(self, username: str) -> Optional[Dict]:
        return self.db.fetch_one(
            "SELECT d.*, u.username FROM drivers d JOIN users u ON d.user_id = u.user_id WHERE u.username = ?",
            (username,)
        )
    
    def accept_order(self, order_id: int, driver_id: int) -> bool:
        responses = self.db.get_order_responses(order_id)
        if responses:
            return False
        
        self.db.add_order_response(order_id, driver_id)
        return True
    
    def is_order_taken(self, order_id: int) -> bool:
        responses = self.db.get_order_responses(order_id)
        return len(responses) > 0
    
    def get_order_info(self, order_id: int) -> Optional[Dict]:
        return self.db.get_order(order_id)
    
    def get_driver_orders_history(self, driver_id: int, limit: int = 10) -> List[Dict]:
        return self.db.get_driver_orders(driver_id, limit)
    
    def get_order_responses_info(self, order_id: int) -> List[Dict]:
        return self.db.get_order_responses(order_id)
    
    def get_all_users(self) -> List[Dict]:
        return self.db.get_all_users()
    
    def get_all_drivers(self) -> List[Dict]:
        return self.db.get_all_drivers()
    
    def export_users(self) -> str:
        users = self.get_all_users()
        if not users:
            return "Нет пользователей для экспорта"
        
        result = "Список пользователей:\n\n"
        for user in users:
            result += f"ID: {user['user_id']}\n"
            result += f"Username: @{user['username']}\n"
            result += f"Имя: {user['first_name']} {user['last_name']}\n"
            result += f"Роль: {user['role']}\n"
            result += f"Дата регистрации: {user['created_at']}\n"
            result += "─" * 30 + "\n"
        
        return result
    
    def export_users_excel(self) -> str:
        users = self.get_all_users()
        if not users:
            return None
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Пользователи"
        
        headers = ["ID", "Username", "Имя", "Фамилия", "Роль", "Дата регистрации"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        for row_num, user in enumerate(users, 2):
            ws.cell(row=row_num, column=1, value=user['user_id'])
            ws.cell(row=row_num, column=2, value=f"@{user['username']}" if user['username'] else "")
            ws.cell(row=row_num, column=3, value=user['first_name'] or "")
            ws.cell(row=row_num, column=4, value=user['last_name'] or "")
            ws.cell(row=row_num, column=5, value=user['role'])
            ws.cell(row=row_num, column=6, value=user['created_at'])
        
        column_widths = [10, 15, 15, 15, 10, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=len(users)+1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
        
        filename = f"users_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(os.getcwd(), filename)
        wb.save(filepath)
        
        return filepath
    
    def export_drivers_excel(self) -> str:
        drivers = self.get_all_drivers()
        if not drivers:
            return None
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Водители"
        
        headers = ["ID", "Username", "ФИО", "Телефон", "Группа", "Дата регистрации"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        for row_num, driver in enumerate(drivers, 2):
            group_name = "Не указана"
            if driver.get('group_id'):
                group = self.db.get_group(driver['group_id'])
                if group:
                    group_name = group['group_name']
            
            ws.cell(row=row_num, column=1, value=driver['user_id'])
            ws.cell(row=row_num, column=2, value=f"@{driver['username']}" if driver['username'] else "")
            ws.cell(row=row_num, column=3, value=driver['full_name'])
            ws.cell(row=row_num, column=4, value=driver['phone'])
            ws.cell(row=row_num, column=5, value=group_name)
            ws.cell(row=row_num, column=6, value=driver['created_at'])
        
        column_widths = [10, 15, 25, 15, 15, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=len(drivers)+1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
        
        filename = f"drivers_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(os.getcwd(), filename)
        wb.save(filepath)
        
        return filepath
    
    def export_drivers(self) -> str:
        drivers = self.get_all_drivers()
        if not drivers:
            return "Нет водителей для экспорта"
        
        result = "Список водителей:\n\n"
        for driver in drivers:
            group_name = "Не указана"
            if driver.get('group_id'):
                group = self.db.get_group(driver['group_id'])
                if group:
                    group_name = group['group_name']
            
            result += f"ID: {driver['user_id']}\n"
            result += f"ФИО: {driver['full_name']}\n"
            result += f"Телефон: {driver['phone']}\n"
            result += f"Группа: {group_name}\n"
            result += f"Username: @{driver['username']}\n"
            result += "─" * 30 + "\n"
        
        return result
    
    def add_group(self, group_name: str) -> int:
        return self.db.add_group(group_name)

    def get_all_groups(self) -> List[Dict]:
        return self.db.get_all_groups()

    def delete_group(self, group_id: int):
        self.db.delete_group(group_id)

    def get_group_by_name(self, group_name: str) -> Optional[Dict]:
        return self.db.get_group_by_name(group_name)

    def register_driver(self, user_id: int, full_name: str, phone: str, group_id: int):
        self.db.add_driver(user_id, full_name, phone, group_id)

    def get_drivers_by_group(self, group_id: int) -> List[Dict]:
        return self.db.get_drivers_by_group(group_id)

    def create_order_with_topic(self, admin_id: int, description: str, group_id: int, photos: List[str], topic_name: str) -> int:
        return self.db.add_order(admin_id, description, group_id, photos, topic_name)

    def add_driver_offer(self, order_id: int, driver_id: int, price: float, comment: str = None):
        self.db.add_driver_offer(order_id, driver_id, price, comment)

    def get_order_offers(self, order_id: int) -> List[Dict]:
        return self.db.get_order_offers(order_id)

    def accept_driver_offer(self, offer_id: int) -> bool:
        return self.db.accept_driver_offer(offer_id)